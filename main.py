import openpyxl
import argparse
import sys
import os
import datetime
import ipaddress
from openpyxl.worksheet.table import Table, TableStyleInfo
from tabulate import tabulate
from openpyxl.worksheet.table import Table, TableStyleInfo



def col_num_to_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def parse_firewall_config(config_text):
    original_dict = {
        "Line": "",
        "Name": "",
        "Enabled": "",
        "Action": "",
        "IP Version": "",
        "Source CIDR": "",
        "Destination CIDR": "",
        "Group": "",
        "Protocol": "",
        "Destination Port": "",
        "Remote Port": "",
        "VLAN": "",
        "Address ID": "",
        "Return Source MAC": "",
        "Allow Return": "",
    }
    filter_data = {}
    filter_list = []
    filter_id = 0
    sip = ""
    sip_cidr = ""
    dip = ""
    dip_cidr = ""
    x = 0

    for line in config_text:
        stripped_line = line.strip()

        if stripped_line.startswith('/c/slb/filt'):
            filter_id = stripped_line.split()[-1]
            if not filter_id.isnumeric():
                # print("Not a number")
                filter_id = int(filter_id.split('/')[0])
                if len(filter_data) == 0:
                    filter_data = original_dict.copy()
                    filter_data["Line"] = filter_id
                elif filter_id != filter_data["Line"]:
                    print(filter_data)
                    filter_list.append(filter_data)
                    filter_data = original_dict.copy()
                    filter_data["Line"] = filter_id
            else:
                filter_id = int(filter_id)
                if int(filter_id) == 1:
                    x = 1
                    filter_data = original_dict.copy()
                    filter_data["Line"] = filter_id
                else:
                    print(filter_data)
                    filter_list.append(filter_data)
                    filter_data = original_dict.copy()
                    filter_data["Line"] = filter_id
            # if filter_id != filter_item["Line"]:
            #     filter_item = original_dict.copy()
            #     filter_item["Line"] = filter_id

        elif x == 1 and stripped_line.startswith('/') and not stripped_line.startswith('/c/slb/filt'):
            break

        elif filter_data and stripped_line:
            key_value = stripped_line.split(maxsplit=1)
            if key_value[0] == "ena" or key_value[0] == "dis":
                value = key_value[0]
                key_value[0] = "Enabled"
                if value == "ena":
                    key_value.append("Yes")
                else:
                    key_value.append("No")
            elif key_value[0] == "name":
                key_value[0] = "Name"
                key_value[1] = key_value[1].strip('"')
            elif key_value[0] == "action":
                key_value[0] = "Action"
            elif key_value[0] == "ipver":
                key_value[0] = "IP Version"
            elif key_value[0] == "group":
                key_value[0] = "Group"
            elif key_value[0] == "sip":
                if key_value[1] == "any":
                    key_value[1] = "0.0.0.0"
                sip = key_value[1]
                continue
            elif key_value[0] == "smask":
                print(sip, " ", dip)
                if key_value[0] == "0.0.0.0":
                    print("Stop")
                key_value[0] = "Source CIDR"
                mask = key_value[1]
                key_value[1] = drop_host_bits(sip, mask)
            elif key_value[0] == "dip":
                if key_value[1] == "any":
                    key_value[1] = "0.0.0.0"
                dip = key_value[1]
                continue
            elif key_value[0] == "dmask":
                key_value[0] = "Destination CIDR"
                mask = key_value[1]
                key_value[1] = drop_host_bits(dip, mask)
            elif key_value[0] == "rport":
                key_value[0] = "Remote Port"
            elif key_value[0] == "proto":
                key_value[0] = "Protocol"
            elif key_value[0] == "dport":
                key_value[0] = "Destination Port"
            elif key_value[0] == "vlan":
                key_value[0] = "VLAN"
            elif key_value[0] == "add":
                key_value[0] = "Address ID"
            elif key_value[0] == "rtsrcmac":
                key_value[0] = "Return Source MAC"
                key_value[1] = "yes"
            elif key_value[0] == "reverse":
                key_value[0] = "Allow Return"
                key_value[1] = "yes"

            if len(key_value) == 2:
                key, value = key_value
                filter_data[key] = value

    # Add the last NAT configuration if it exists
    if filter_data:
        filter_list.append(filter_data)

    return filter_list


def parse_nat_config(config_text):

    nat_list = []
    original_dict = {
        "Line": "",
        "Rule": "",
        "IP Version": "",
        "Wan Link": "",
        "Local Address": "",
        "NAT Address": "",
        "Type": "",
        "No NAT": "",
        "Name": "",
    }
    nat_data = {}
    line_number = 1

    for line in config_text:
        stripped_line = line.strip()

        # Check for the beginning of a NAT or NO_NAT configuration
        if stripped_line.startswith('/c/slb/lp/nat'):

            # If we're already capturing a NAT configuration, add it to the list
            if nat_data:
                nat_list.append(nat_data)
                line_number += 1


            # Start capturing a new NAT configuration
            nat_data = original_dict.copy()
            # Write the line number
            nat_data["Line"] = line_number
            # Write the 2nd half of the line to the Name key
            nat_data["Rule"] = stripped_line.split()[-1]

        # Check for the end of a NAT configuration
        elif stripped_line.startswith('/') and not stripped_line.startswith('/c/slb/lp/nat'):
            if nat_data:
                nat_list.append(nat_data)
                nat_data = None

        # If we're within a NAT configuration, capture the relevant information
        elif nat_data and stripped_line:
            key_value = stripped_line.split(maxsplit=2)
            if key_value[0] == "ipver":
                key_value[0] = "IP Version"
            elif key_value[0] == "name":
                key_value[0] = "Name"
                key_value[1] = key_value[1].strip('"')
            elif key_value[0] == "wanlink":
                key_value[0] = "Wan Link"
            elif key_value[0] == "locladd":
                key_value[0] = "Local Address"
            elif key_value[0] == "natadd":
                key_value[0] = "NAT Address"
            elif key_value[0] == "type":
                key_value[0] = "Type"

            if len(key_value) == 2:
                key, value = key_value
                nat_data[key] = value
            elif key_value[0] == "Local Address" or key_value[0] == "NAT Address":
                key, address, mask = key_value
                cidr = drop_host_bits(address, mask)
                nat_data[key] = cidr

    # Add the last NAT configuration if it exists
    if nat_data:
        nat_list.append(nat_data)

    return nat_list


def convert_text_to_numbers(worksheet):
    """
    Convert cells with numbers formatted as text to actual numbers.

    Args:
    - worksheet: The worksheet where the conversion needs to be done.

    Returns:
    - None
    """
    for row in worksheet.iter_rows():
        for cell in row:
            try:
                print(f"{cell.column_letter}:{cell.row}  Data: {cell.value}")
                cell.value = float(cell.value)
            except ValueError:
                # The cell value is not a number, so we skip it.
                pass


def auto_resize_columns(worksheet):
    """
    Resize columns to fit the width of the text.

    Args:
    - worksheet: The worksheet whose columns need to be resized.

    Returns:
    - None
    """
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


def create_and_style_table(workbook, tab_name, table_name="", table_style="Table Style Light 1"):
    """
    Create a table for the data in the specified worksheet and apply a style to it.

    Args:
    - workbook: The workbook containing the worksheet.
    - tab_name: The name of the worksheet/tab to format.
    - table_name: The name of the table to create (default is "DataTable").
    - table_style: The name of the style of table format (default is "White, Table Style Light 1")

    Returns:
    - None
    """


    # Get the worksheet by its name
    worksheet = workbook[tab_name]

    # Determine the range of the data
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    column_number = col_num_to_letter(max_col)
    data_range = f"A1:{column_number}{max_row}"
    print(data_range)

    # Create the table
    table = Table(displayName=table_name, ref=data_range)

    # Apply the desired table style
    style = TableStyleInfo(
        name=table_style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    worksheet.add_table(table)
    auto_resize_columns(worksheet)
    convert_text_to_numbers(worksheet)


def is_in_same_subnet(address, cidr):
    ip = ipaddress.ip_address(address)
    network = ipaddress.ip_network(cidr, strict=False)
    return ip in network


def subnet_mask_to_cidr(mask):
    return sum(bin(int(x)).count('1') for x in mask.split('.'))


def drop_host_bits(ip, mask):
    cidr = subnet_mask_to_cidr(mask)
    ip_parts = list(map(int, ip.split('.')))

    # Calculate the number of host bits
    host_bits = 32 - cidr

    # Zero out the host bits
    for i in range(4):
        if host_bits <= 0:
            break
        ip_parts[3 - i] &= ~((1 << min(8, host_bits)) - 1)
        host_bits -= 8

    return '.'.join(map(str, ip_parts)) + '/' + str(cidr)


def populate_sheet(dataset, sheet):
    # Add headers to the L3 sheet
    # Write headers to the first row
    headers = dataset[0].keys()
    for col_num, header in enumerate(headers, 1):
        col_letter = sheet.cell(row=1, column=col_num).column_letter
        sheet[f"{col_letter}1"] = header

    # Write data for each dictionary entry
    for row_num, line in enumerate(dataset, 2):
        for col_num, data in enumerate(line.values(), 1):

            col_letter = sheet.cell(row=row_num, column=col_num).column_letter
            sheet[f"{col_letter}{row_num}"] = data


def parse_l3_data(lines):
    l3_data = []
    original_dict = {
        "Interface": None,
        "Enabled": "",
        "IP Version": "",
        "IPv4 Address": "",
        "Mask": "",
        "CIDR": "",
        "Broadcast": "",
        "Peer": "",
        "VLAN": "",
        "Description": ""
    }
    interface_data = original_dict.copy()
    mgmt_if = {
        "Address": "",
        "Mask": "",
        "Broadcast": "",
        "Gateway": "",
        "CIDR": "",
        "Enabled": "",
        "SNMP_Enabled": "",
        "SYSLOG_Enabled": "",
        "RADIUS_Enabled": "",
        }

    # Track Interface Configuration Sections
    x = 0
    # Track Management Interface Configuration Sections
    y = 0

    for line in lines:
        line = line.strip()
        if y != 2:
            if line.startswith("/c/sys/mmgmt"):
                y = 1
            elif y == 1 and "addr" in line:
                mgmt_if["Address"] = line.split()[-1]
            elif y == 1 and "mask" in line:
                mgmt_if["Mask"] = line.split()[-1]
            elif y == 1 and "broad" in line:
                mgmt_if["Broadcast"] = line.split()[-1]
            elif y == 1 and "gw" in line:
                mgmt_if["Gateway"] = line.split()[-1]
            elif y == 1 and "ena" in line:
                mgmt_if["Enabled"] = "yes"
            elif y == 1 and "snmp" in line:
                mgmt_if["SNMP_Enabled"] = line.split()[-1]
            elif y == 1 and "syslog" in line:
                mgmt_if["SYSLOG_Enabled"] = line.split()[-1]
            elif y == 1 and "radius" in line:
                mgmt_if["RADIUS_Enabled"] = line.split()[-1]
            elif y == 1 and line.startswith("/"):
                y = 2
                if mgmt_if["Enabled"] == "":
                    mgmt_if["Enabled"] = "no"
                mgmt_if["CIDR"] = drop_host_bits(mgmt_if["Address"], mgmt_if["Mask"])
                continue

        if x == 1 and line.startswith("/") and "/c/l3/if" not in line:
            x = 2
            l3_data.append(interface_data)
            interface_data = original_dict.copy()  # Reset interface data
            continue
        elif "/c/l3/if" in line and x != 2:
            if interface_data["Interface"] is not None:  # If there's existing data, append it to the list
                l3_data.append(interface_data)
                interface_data = original_dict.copy()  # Reset for the next interface
            interface_data["Interface"] = line.split()[-1]
            x = 1
        elif x == 2:
            continue
        elif "ena" in line and x == 1:
            # interface_data["enabled"] = line.split()[-1]
            interface_data["Enabled"] = "yes"
        elif "ipver" in line and x == 1:
            interface_data["IP Version"] = line.split()[-1]
        elif "addr" in line and x == 1:
            interface_data["IPv4 Address"] = line.split()[-1]
            # Copy values from mgmt_if if IP address matches
            if is_in_same_subnet(interface_data["IPv4 Address"], mgmt_if["CIDR"]):
                print(f'Interface is currently: {interface_data["IPv4 Address"]} \nMgmt CIDR: {mgmt_if["CIDR"]}')
                interface_data["CIDR"] = mgmt_if["CIDR"]
                interface_data["Broadcast"] = mgmt_if["Broadcast"]
                interface_data["Mask"] = mgmt_if["Mask"]
        elif "mask" in line and x == 1:
            interface_data["Mask"] = line.split()[-1]
            interface_data["CIDR"] = drop_host_bits(interface_data["IPv4 Address"], interface_data["Mask"])
        elif "broad" in line and x == 1:
            interface_data["Broadcast"] = line.split()[-1]
        elif "vlan" in line and x == 1:
            interface_data["VLAN"] = line.split()[-1]
        elif "peer" in line and x == 1:
            interface_data["Peer"] = line.split()[-1]
        elif "descr" in line and x == 1:
            interface_data["Description"] = line.split('"')[1]

    if interface_data and interface_data["Interface"] is not None:  # Append the last interface's data
        l3_data.append(interface_data)

    return l3_data

def parse_l3_data_2(lines):
    l3_data = []
    original_dict = {
        "Interface": None,
        "Enabled": "no",
        "IP Version": "",
        "IPv4 Address": "",
        "Mask": "",
        "CIDR": "",
        "Broadcast": "",
        "Peer": "",
        "VLAN": "",
        "Description": ""
    }
    interface_data = original_dict.copy()
    mgmt_if = {
        "Address": "",
        "Mask": "",
        "Broadcast": "",
        "Gateway": "",
        "CIDR": "",
        "Enabled": "no",
        "SNMP_Enabled": "",
        "SYSLOG_Enabled": "",
        "RADIUS_Enabled": "",
    }

    mgmt_mapping = {
        "addr": "Address",
        "mask": "Mask",
        "broad": "Broadcast",
        "gw": "Gateway",
        "ena": "Enabled",
        "snmp": "SNMP_Enabled",
        "syslog": "SYSLOG_Enabled",
        "radius": "RADIUS_Enabled"
    }

    l3_mapping = {
        "ena": "Enabled",
        "ipver": "IP Version",
        "addr": "IPv4 Address",
        "mask": "Mask",
        "broad": "Broadcast",
        "vlan": "VLAN",
        "peer": "Peer",
        "descr": "Description"
    }

    for line in lines:
        line = line.strip()

        if line.startswith("/c/sys/mmgmt"):
            for line in lines:
                line = line.strip()
                key = line.split()[0]
                if key in mgmt_mapping:
                    mgmt_if[mgmt_mapping[key]] = line.split()[-1]
                elif line.startswith("/"):
                    break

        elif line.startswith("/c/l3/if"):
            if interface_data["Interface"] is not None:
                l3_data.append(interface_data)
                interface_data = original_dict.copy()
            interface_data["Interface"] = line.split()[-1]
            for line in lines:
                line = line.strip()
                key = line.split()[0]
                if key in l3_mapping:
                    if key == "descr":
                        interface_data[l3_mapping[key]] = line.split('"')[1]
                    else:
                        interface_data[l3_mapping[key]] = line.split()[-1]
                elif line.startswith("/"):
                    break

    if interface_data["Interface"] is not None:
        l3_data.append(interface_data)

    return l3_data


def create_parse_populate_style(workbook, payload, sheet_name, parse_string="", table_style="TableStyleMedium9"):
    sheet = workbook.create_sheet(sheet_name)
    populate_sheet(payload, sheet)
    table_name = sheet_name.replace(" ", "_")
    create_and_style_table(workbook, sheet_name, table_name=table_name, table_style=table_style)

def parse_gslb_config2(lines):
    gslb_rules = []
    original_dict = {
        "Rule ID": None,
        "Enabled": "",
        "Name": "",
        "Type": "",
        "TTL": "",
        "RR": "",
        "DName": "",
        "Fallback": "",
        "Metrics": []
    }
    metric_data = {
        "Metric ID": None,
        "GMetric": "",
        "AddNet": ""
    }

    # Track GSLB Rule Configuration Sections
    x = 0

    for line in lines:
        line = line.strip()

        if line.startswith("/c/slb/gslb/rule"):
            rule_id = line.split()[-1]
            if not rule_id.isnumeric():
                if "metric" in line:
                    metric_data["Metric ID"] = line.split()[-1]
                    x = 2
                else:
                    continue
            else:
                if rule_data["Rule ID"] is not None:  # If there's existing data, append it to the list
                    gslb_rules.append(rule_data)
                    rule_data = original_dict.copy()  # Reset for the next rule
                rule_data["Rule ID"] = rule_id
                x = 1
        elif x == 2 and line.startswith("/"):
            rule_data["Metrics"].append(metric_data)
            metric_data = {
                "Metric ID": None,
                "GMetric": "",
                "AddNet": ""
            }
            x = 1
            continue
        elif "ena" in line and x == 1:
            rule_data["Enabled"] = "yes"
        elif "name" in line and x == 1:
            rule_data["Name"] = line.split('"')[1]
        elif "type" in line and x == 1:
            rule_data["Type"] = line.split()[-2]
        elif "ttl" in line and x == 1:
            rule_data["TTL"] = line.split()[-1]
        elif "rr" in line and x == 1:
            rule_data["RR"] = line.split()[-1]
        elif "dname" in line and x == 1:
            rule_data["DName"] = line.split('"')[1]
        elif "fallback" in line and x == 1:
            rule_data["Fallback"] = "yes"
        elif "gmetric" in line and x == 2:
            metric_data["GMetric"] = line.split()[-1]
        elif "addnet" in line and x == 2:
            metric_data["AddNet"] = line.split()[-1]

    if rule_data and rule_data["Rule ID"] is not None:  # Append the last rule's data
        gslb_rules.append(rule_data)

    return gslb_rules


def parse_gslb_network_config(config_text, search_string, gslb_rules_list):


    original_dict = {
        "Network": "",
        "Enabled": "",
        "Server Type": "",
        "Server IP": "",
        "WAN Group": "",
    }
    gslb_network_id = ""
    gslb_network_data = {}
    gslb_network_list = []
    x = 0

    for line in config_text:
        stripped_line = line.strip()
        if stripped_line.startswith(search_string):
            rule_start = stripped_line.split()
            gslb_network_id = rule_start[1]
            gslb_network_id = int(gslb_network_id)
            if x == 0:
                x += 1
                gslb_network_data = original_dict.copy()
                gslb_network_data["Network"] = gslb_network_id
            else:
                print(gslb_network_data)
                gslb_network_list.append(gslb_network_data)
                gslb_network_data = original_dict.copy()
                gslb_network_data["Network"] = gslb_network_id


        elif x == 1 and stripped_line.startswith('/') and not stripped_line.startswith(search_string):
            break

        elif gslb_network_data and stripped_line:
            key_value = stripped_line.split(maxsplit=1)
            if key_value[0] == "ena" or key_value[0] == "dis":
                value = key_value[0]
                key_value[0] = "Enabled"
                if value == "ena":
                    key_value.append("Yes")
                else:
                    key_value.append("No")
            elif key_value[0] == "servtyp":
                key_value[0] = "Server Type"
            elif key_value[0] == "servip":
                key_value[0] = "Server IP"
            elif key_value[0] == "wangrp":
                key_value[0] = "WAN Group"

            if len(key_value) == 2:
                key, value = key_value
                gslb_network_data[key] = value

    # Add the last NAT configuration if it exists
    if gslb_network_data:
        gslb_network_list.append(gslb_network_data)

    return gslb_network_list


def combine_gslb_data(gslb_rules, gslb_networks):
    import copy
    new_rules_data = copy.deepcopy(gslb_rules)
    # Track the network number being associated
    addnet = ""
    # Track line in gslb_rules
    x = 0
    # Track line in gslb_networks
    y = 0

    for lines in gslb_rules:
        if addnet == "":
            pass
        else:
            x += 1
        for key, value in lines.items():
            if key == "Metric 1 addnet":
                addnet = int(value)
                for network in gslb_networks:
                    for net_key, net_value in network.items():
                        if net_key == "Network":
                            if int(value) == int(net_value):
                                new_rules_data[x].update({
                                    "Ntwk Enabled": network["Enabled"],
                                    "Server Type": network["Server Type"],
                                    "Server IP": network["Server IP"],
                                    "WAN Group": network["WAN Group"]
                                })
                                break

    return new_rules_data



def parse_gslb_rules_config(config_text, search_string):
    original_dict = {
        "Rule": "",
        "Name": "",
        "Enabled": "",
        "Type": "",
        "TTL": "",
        "RR": "",
        "DNS Name": "",
        "Fallback": "",
        "Metric 1 gmetric": "",
        "Metric 1 addnet": "",
        "Metric 3 gmetric": "",
        "Metric 3 addnet": "",
    }
    gslb_rule_data = {}
    gslb_rule_list = []
    x = 0
    metric = ""

    for line in config_text:
        stripped_line = line.strip()

        if stripped_line.startswith(search_string):
            rule_start = stripped_line.split()
            metric = ""
            if len(rule_start) == 2:
                gslb_rule_id = rule_start[1]
            elif len(rule_start) == 3:
                metric = rule_start[2]
                key = rule_start[1].split("/")[1]
                stripped_line = [key, metric]
                continue
            if not gslb_rule_id.isnumeric():
                print("Not a number")
                gslb_rule_id = int(gslb_rule_id.split('/')[0])
                if len(gslb_rule_data) == 0:
                    gslb_rule_data = original_dict.copy()
                    gslb_rule_data["Rule"] = gslb_rule_id
                elif gslb_rule_id != gslb_rule_data["Rule"]:
                    x += 1
                    print(gslb_rule_data)
                    gslb_rule_list.append(gslb_rule_data)
                    gslb_rule_data = original_dict.copy()
                    gslb_rule_data["Rule"] = gslb_rule_id
            else:
                gslb_rule_id = int(gslb_rule_id)
                if x == 0:
                    x += 1
                    gslb_rule_data = original_dict.copy()
                    gslb_rule_data["Rule"] = gslb_rule_id
                else:
                    print(gslb_rule_data)
                    gslb_rule_list.append(gslb_rule_data)
                    gslb_rule_data = original_dict.copy()
                    gslb_rule_data["Rule"] = gslb_rule_id

        elif x == 1 and stripped_line.startswith('/') and not stripped_line.startswith(search_string):
            break

        elif gslb_rule_data and stripped_line:
            key_value = stripped_line.split(maxsplit=1)
            if key_value[0] == "ena" or key_value[0] == "dis":
                value = key_value[0]
                key_value[0] = "Enabled"
                if value == "ena":
                    key_value.append("Yes")
                else:
                    key_value.append("No")
            elif key_value[0] == "name":
                key_value[0] = "Name"
                key_value[1] = key_value[1].strip('"')
            elif key_value[0] == "type":
                key_value[0] = "Type"
            elif key_value[0] == "ttl":
                key_value[0] = "TTL"
            elif key_value[0] == "rr":
                key_value[0] = "RR"
            elif key_value[0] == "dname":
                key_value[0] = "DNS Name"
                key_value[1] = key_value[1].replace('"', '')
            elif key_value[0] == "fallback":
                if key_value[0] == "ena":
                    key_value.append("Yes")
                else:
                    key_value.append("No")
                key_value[0] = "Fallback"
            elif key_value[0] == "gmetric":
                key_value[0] = "Metric " + str(metric) + " gmetric"
            elif key_value[0] == "addnet":
                key_value[0] = "Metric " + str(metric) + " addnet"

            if len(key_value) == 2:
                key, value = key_value
                gslb_rule_data[key] = value

    # Add the last NAT configuration if it exists
    if gslb_rule_data:
        gslb_rule_list.append(gslb_rule_data)

    # Run parser for the network configuration to add to the dataset
    gslb_networks = parse_gslb_network_config(config_text, "/c/slb/gslb/network", gslb_rule_list)
    # Combine data into the gslb_rules
    combined_list = combine_gslb_data(gslb_rule_list, gslb_networks)

    return combined_list


def create_excel_output(lines, output_path):
    wb = openpyxl.Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    l3_sheet_name = "Layer 3"
    nat_sheet_name = "NAT"
    firewall_sheet_name = "Firewall"
    gslb_network_sheet_name = "GSLB Network"
    gslb_rules_sheet_name = "GSLB Rules"

    l2_sheet = wb.create_sheet("Layer 2")
    # gslb_network_sheet = wb.create_sheet(gslb_network_sheet_name)
    # gslb_rule_sheet = wb.create_sheet(gslb_rules_sheet_name)
    # nat_sheet = wb.create_sheet(nat_sheet_name)
    # firewall_sheet = wb.create_sheet(firewall_sheet_name)


    # Parse config for L3 data and create Excel tab formatted with output
    l3_dataset = parse_l3_data(lines)
    create_parse_populate_style(wb, l3_dataset, l3_sheet_name, table_style="TableStyleMedium11")
    # Parse config for L3 data and create Excel tab formatted with output
    nat_dataset = parse_nat_config(lines)
    create_parse_populate_style(wb, nat_dataset, nat_sheet_name, table_style="TableStyleMedium9")
    # Parse config for L3 data and create Excel tab formatted with output
    firewall_dataset = parse_firewall_config(lines)
    create_parse_populate_style(wb, firewall_dataset, firewall_sheet_name, table_style="TableStyleMedium10")
    # Parse config for GSLB Rules and create Excel tab formatted with output
    gslb_rules_dataset = parse_gslb_rules_config(lines, "/c/slb/gslb/rule")
    tabulate(gslb_rules_dataset)
    create_parse_populate_style(wb, gslb_rules_dataset, gslb_rules_sheet_name, table_style="TableStyleMedium8")
    # Generic Sheets
    # populate_sheet_basic(l2_sheet, lines, "/c/l2/")
    # populate_sheet_basic(summary_sheet, lines, "/c/sys/")
    # populate_sheet_basic(gslb_network_sheet_name, lines, "/c/slb/gslb/network")
    # populate_sheet_basic(gslb_rules_sheet_name, lines, "/c/slb/gslb/rule")
    # nat_dataset = parse_nat_config(lines)
    # populate_sheet(nat_dataset, nat_sheet)
    # create_and_style_table(wb, nat_sheet_name, table_name="NAT_Data", table_style="TableStyleMedium9")
    # firewall_dataset = parse_filter_config(lines)
    # populate_sheet(firewall_dataset, firewall_sheet)
    # create_and_style_table(wb, firewall_sheet_name, table_name="Firewall_Data", table_style="TableStyleMedium9")

    wb.save(output_path)


def populate_sheet_basic(sheet, lines, keyword):
    for line in lines:
        if keyword in line:
            sheet.append([line.strip()])


def parse_file(file_path):
    # Check for argument or environment variable for the file path
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    elif 'CONFIG_FILE_PATH' in os.environ:
        file_path = os.environ['CONFIG_FILE_PATH']
    else:
        print("Please provide the file path as an argument or set the CONFIG_FILE_PATH environment variable.")
        sys.exit(1)

    # Read the file
    with open(file_path, 'r') as f:
        data = f.read()

    # Split the data into lines
    lines = data.strip().split("\n")

    # Return the output of the lines
    return lines


def get_args():
    parser = argparse.ArgumentParser(description="Parse a text file and output to Excel using OpenPyXL.")
    parser.add_argument("file_path", type=str, help="Path to the input text file.")
    parser.add_argument("--output", type=str, default="output.xlsx",
                        help="Path to save the output Excel file. Default is 'output.xlsx'.")

    args = parser.parse_args()

    if not os.path.exists(args.file_path):
        print(f"Error: File '{args.file_path}' does not exist.")
        exit(1)

    return args


def get_standard_filename(input_args):
    # Generate a timestamp
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    dest_filename = args.output
    if dest_filename.endswith(".xlsx"):
        stripped_filename = dest_filename[:-5]
        dest_filename = f"{stripped_filename}_{timestamp}.xlsx"
    elif dest_filename.endswith(".xls"):
        stripped_filename = dest_filename[:-4]
        dest_filename = f"{stripped_filename}_{timestamp}.xlsx"
    else:
        dest_filename = f"{dest_filename}_{timestamp}.xlsx"

    return dest_filename


if __name__ == "__main__":
    # Get arguments from the command line
    args = get_args()

    # Standardize filename output with timestamp
    output_filename = get_standard_filename(args)

    # Parse input file for Summary, L2, L3, GLB and other data
    input_file = parse_file(args.file_path)
    create_excel_output(input_file, output_filename)
    parsed_data = parse_nat_config(input_file)
    # Using tabulate to print the parsed data
    print(tabulate(parsed_data, headers="keys", tablefmt="grid"))
    filter_data = parse_firewall_config(input_file)
    # Using tabulate to print the parsed data
    print(tabulate(filter_data, headers="keys", tablefmt="grid"))
