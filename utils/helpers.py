# Helper functions for this application.
import ipaddress
from openpyxl.worksheet.table import Table, TableStyleInfo


def col_num_to_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def convert_text_to_numbers(worksheet):
    """
    Convert cells with numbers formatted as text to actual numbers.

    Args:
    - worksheet: The worksheet where the conversion needs to be done.

    Returns:
    - None
    """
    for row in worksheet.iter_rows():
        for cell in row:
            try:
                print(f"{cell.column_letter}:{cell.row}  Data: {cell.value}")
                cell.value = float(cell.value)
            except ValueError:
                # The cell value is not a number, so we skip it.
                pass


def auto_resize_columns(worksheet):
    """
    Resize columns to fit the width of the text.

    Args:
    - worksheet: The worksheet whose columns need to be resized.

    Returns:
    - None
    """
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


def create_and_style_table(workbook, tab_name, table_name="", table_style="Table Style Light 1"):
    """
    Create a table for the data in the specified worksheet and apply a style to it.

    Args:
    - workbook: The workbook containing the worksheet.
    - tab_name: The name of the worksheet/tab to format.
    - table_name: The name of the table to create (default is "DataTable").
    - table_style: The name of the style of table format (default is "White, Table Style Light 1")

    Returns:
    - None
    """


    # Get the worksheet by its name
    worksheet = workbook[tab_name]

    # Determine the range of the data
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    column_number = col_num_to_letter(max_col)
    data_range = f"A1:{column_number}{max_row}"
    print(data_range)

    # Create the table
    table = Table(displayName=table_name, ref=data_range)

    # Apply the desired table style
    style = TableStyleInfo(
        name=table_style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    worksheet.add_table(table)
    auto_resize_columns(worksheet)
    convert_text_to_numbers(worksheet)


def is_in_same_subnet(address, cidr):
    ip = ipaddress.ip_address(address)
    network = ipaddress.ip_network(cidr, strict=False)
    return ip in network


def subnet_mask_to_cidr(mask):
    return sum(bin(int(x)).count('1') for x in mask.split('.'))


def drop_host_bits(ip, mask):
    cidr = subnet_mask_to_cidr(mask)
    ip_parts = list(map(int, ip.split('.')))

    # Calculate the number of host bits
    host_bits = 32 - cidr

    # Zero out the host bits
    for i in range(4):
        if host_bits <= 0:
            break
        ip_parts[3 - i] &= ~((1 << min(8, host_bits)) - 1)
        host_bits -= 8

    return '.'.join(map(str, ip_parts)) + '/' + str(cidr)



def combine_gslb_data(gslb_rules, gslb_networks):
    import copy
    new_rules_data = copy.deepcopy(gslb_rules)
    # Track the network number being associated
    addnet = ""
    # Track line in gslb_rules
    x = 0
    # Track line in gslb_networks
    y = 0

    for lines in gslb_rules:
        if addnet == "":
            pass
        else:
            x += 1
        for key, value in lines.items():
            if key == "Metric 1 addnet":
                addnet = int(value)
                for network in gslb_networks:
                    for net_key, net_value in network.items():
                        if net_key == "Network":
                            if int(value) == int(net_value):
                                new_rules_data[x].update({
                                    "Ntwk Enabled": network["Enabled"],
                                    "Server Type": network["Server Type"],
                                    "Server IP": network["Server IP"],
                                    "WAN Group": network["WAN Group"]
                                })
                                break

    return new_rules_data
